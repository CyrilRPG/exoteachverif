# app.py — Vérification I3/I4+ + Export Excel/PDF (1 onglet/page = 1 classe)
# FIXES :
# - Désactivation du file watcher Streamlit pour éviter "inotify instance limit reached"
# - Remplacement de use_container_width=True par width="stretch" (API moderne)
# MODIFS conservées :
# - Ajout colonne "ID" (issue de l'Excel) dans les exports Excel et PDF
# - Suppression de la colonne "Fiches récupérées ?"
# - Tout le reste inchangé

# ==== IMPORTANT : désactiver le watcher AVANT d'importer streamlit ====
import os
os.environ.setdefault("STREAMLIT_SERVER_FILE_WATCHER_TYPE", "none")

import json
import re
from typing import List, Tuple, Dict, Any, Optional, Set
from collections import defaultdict
import io
import unicodedata
from datetime import datetime

import pandas as pd

import streamlit as st
st.set_option("server.fileWatcherType", "none")  # ceinture + bretelles

# ====== PDF (reportlab) ======
try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

# ------------------------------- UI / THEME -------------------------------
st.set_page_config(page_title="Vérif Groupes Étudiants — I3/I4+ & Excel multi-onglets", page_icon="✅", layout="wide")
st.markdown("""
<style>
:root { --radius: 14px; }
.block-container { padding-top: 1rem; }
.stButton>button, .stDownloadButton>button { border-radius: var(--radius); padding:0.55rem 0.9rem; }
.kpi { border:1px solid #e5e7eb; border-radius: var(--radius); padding:0.8rem; background:#fafafa; }
small.dim { color:#6b7280; }
</style>
""", unsafe_allow_html=True)

st.title("Vérification des groupes étudiants — format I3/I4+ & Export Excel par classe")

# ==================== RÉFÉRENTIEL (FILIERES ↔ CLASSES) ====================
FILIERE_NAMES: Dict[int, str] = {
    # USPN
    5016: "LAS - USPN 25/26",
    5017: "PASS - USPN 25/26",
    5018: "LSPS - USPN 25/26",
    # UPC
    5012: "PASS - UPC 25/26",
    5013: "LAS - UPC 25/26",
    # SU
    5014: "PASS - SU (TC) 25/26",
    # UVSQ
    5015: "PASS - UVSQ 25/26",
    # UPS
    5019: "PASS - UPS 25/26",
    # UPEC
    5020: "LAS1 Majeure disciplinaire - UPEC 25/26",
    5021: "LSPS1 - UPEC 25-26",
    5022: "LSPS2 - UPEC 25-26",
    5032: "LSPS3 - UPEC - 25-26",
    # PAES
    5023: "PAES - Présentiel 25-26",
    5024: "PAES - Distanciel 25-26",
    # Terminale Santé
    5025: "Terminale Santé 25-26 - Présentiel",
    5026: "Terminale Santé 25-26 - Distanciel",
    # Première Élite
    5027: "Première Élite 25-26",
}

CLASS_NAMES: Dict[int, str] = {
    # USPN
    5944: "USPN - Classe 1 (LAS) 25/26",
    5943: "USPN - Classe 2 (PASS/LSPS) 25/26",
    5942: "USPN - Classe 1 (PASS/LSPS) 25/26",
    # UPC (PASS)
    5935: "PASS UPC - Classe 4 25/26",
    5934: "PASS UPC - Classe 3 25/26",
    5933: "PASS UPC - Classe 2 25/26",
    5932: "PASS UPC - Classe 1 25/26",
    # UPC (LAS)
    5931: "LAS UPC - Classe 1 25/26",
    # SU
    5940: "PASS SU - Classe 5 (Mineure Sciences) 25/26",
    5939: "PASS SU - Classe 4 (Mineure Lettres) 25/26",
    5938: "PASS SU - Classe 3 (Mineure Sciences) 25/26",
    5937: "PASS SU - Classe 2 (Mineure Sciences) 25/26",
    5936: "PASS SU - Classe 1 (Mineure Sciences) 25/26",
    # UVSQ
    5941: "PASS UVSQ - Classe 1 25/26",
    # UPS
    5945: "PASS UPS - Classe 1 25/26",
    # UPEC
    5953: "LSPS2 UPEC - Classe 3 (25-26)",
    5952: "LSPS2 UPEC - Classe 2 (25-26)",
    5951: "LSPS2 UPEC - Classe 1 (25-26)",
    5950: "LSPS1 UPEC - Classe 4 25/26",
    5949: "LSPS1 UPEC - Classe 3 25/26",
    5948: "LSPS1 UPEC - Classe 2 25/26",
    5947: "LSPS1 UPEC - Classe 1 25-26",
    5946: "LAS1 Majeure disciplinaire - UPEC - Classe 1 25/26",
    # PAES
    6127: "PAES Distanciel - Classe 1 25/26",
    6125: "PAES Présentiel - Classe 4 25/26",
    6124: "PAES Présentiel - Classe 2 25/26",
    6123: "PAES Présentiel - Classe 3 25/26",
    6122: "PAES Présentiel - Classe 1 25/26",
    # Terminale Santé
    6120: "Terminale Santé Distanciel - Classe 1 25/26",
    6119: "Terminale Santé Présentiel - Classe 8 25/26",
    6118: "Terminale Santé Présentiel - Classe 7 25/26",
    6117: "Terminale Santé Présentiel - Classe 6 25/26",
    6116: "Terminale Santé Présentiel - Classe 5 25/26",
    6115: "Terminale Santé Présentiel - Classe 4 25/26",
    6114: "Terminale Santé Présentiel - Classe 3 25/26",
    6113: "Terminale Santé Présentiel - Classe 2 25/26",
    6112: "Terminale Santé Présentiel - Classe 1 25/26",
    # Première Élite
    6128: "Première Elite - Classe 1 25/26",
}

# FILIERE -> CLASSES autorisées
FILIERE_TO_CLASSES: Dict[int, Set[int]] = {
    5016: {5944},
    5017: {5942, 5943},
    5018: {5942, 5943},
    5012: {5932, 5933, 5934, 5935},
    5013: {5931},
    5014: {5936, 5937, 5938, 5939, 5940},
    5015: {5941},
    5019: {5945},
    5020: {5946},
    5021: {5947, 5948, 5949, 5950},
    5022: {5951, 5952, 5953},
    5032: set(),
    5023: {6122, 6123, 6124, 6125},
    5024: {6127},
    5025: {6112, 6113, 6114, 6115, 6116, 6117, 6118, 6119},
    5026: {6120},
    5027: {6128},
}

# Inverse : CLASSE -> FILIERES
CLASSES_TO_FILIERES: Dict[int, Set[int]] = defaultdict(set)
for fcode, cls_set in FILIERE_TO_CLASSES.items():
    for c in cls_set:
        CLASSES_TO_FILIERES[c].add(fcode)

# OFFICIEL (tous codes) pour la détection
OFFICIEL: Dict[int, Tuple[str, str]] = {}
for f_code, f_name in FILIERE_NAMES.items():
    OFFICIEL[f_code] = (f_name, "Filière")
for c_code, c_name in CLASS_NAMES.items():
    OFFICIEL[c_code] = (c_name, "Classe")

# Exceptions : OK si classe seule (vérif) + EXCLUS de l'Excel/PDF
EXCEPTION_OK_IF_CLASS_ONLY: Set[int] = {
    4538, 4537, 4388, 4386, 4385, 4384, 4383, 4382, 4381, 4380, 4379, 4378, 4377, 4376, 4375
}

NUM_RE = re.compile(r"\d+")

def parse_numeros(groupes_str: Any) -> List[int]:
    if pd.isna(groupes_str):
        return []
    return [int(m.group(0)) for m in NUM_RE.finditer(str(groupes_str))]

# ====== helpers : exclusion "Salomé Galbois" (sans accent/casse) ======
def _normalize(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def is_salome_galbois(nom: str, prenom: str) -> bool:
    return _normalize(nom) == "galbois" and _normalize(prenom) == "salome"

# ============================= ANALYSE =============================
def analyser_groupes(groupes_str: Any) -> str:
    nums = parse_numeros(groupes_str)
    has_exception = any(n in EXCEPTION_OK_IF_CLASS_ONLY for n in nums)

    filieres = [n for n in nums if n in OFFICIEL and OFFICIEL[n][1] == "Filière"]
    classes  = [n for n in nums if n in OFFICIEL and OFFICIEL[n][1] == "Classe"]

    if len(filieres) == 0 and len(classes) == 0:
        return "Pas de classe ni de filière"

    if len(filieres) == 0 and len(classes) > 0:
        if has_exception:
            return "OK"
        return "Pas de filière"

    if len(classes) == 0 and len(filieres) > 0:
        return "Pas de classe"

    if len(filieres) > 1 and len(classes) > 1:
        return "Plusieurs filières et plusieurs classes"
    if len(filieres) > 1:
        return "Plusieurs filières"
    if len(classes) > 1:
        return "Plusieurs classes"

    f = filieres[0]
    c = classes[0]
    if c in CLASSES_TO_FILIERES and f in CLASSES_TO_FILIERES[c]:
        return "OK"
    else:
        return "Classe et filière incohérents"

def extra_info(groupes_str: Any) -> Dict[str, Any]:
    nums = parse_numeros(groupes_str)
    connus = [n for n in nums if n in OFFICIEL]
    inconnus = [n for n in nums if n not in OFFICIEL]
    filieres = [n for n in nums if n in OFFICIEL and OFFICIEL[n][1]=="Filière"]
    classes  = [n for n in nums if n in OFFICIEL and OFFICIEL[n][1]=="Classe"]
    filiere_label = FILIERE_NAMES[filieres[0]] if len(filieres)==1 and filieres[0] in FILIERE_NAMES else None
    classe_label = CLASS_NAMES[classes[0]] if len(classes)==1 and classes[0] in CLASS_NAMES else None
    return {
        "NumerosTrouvés": nums,
        "NumerosConnus": connus,
        "NumerosInconnus": inconnus,
        "FiliereDéduite": filiere_label,
        "ClasseDéduite": classe_label,
    }

# ============================= IMPORT I3/I4+ =============================
def excel_col_to_index(col_letter: str) -> int:
    col_letter = col_letter.strip().upper()
    total = 0
    for ch in col_letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError("Lettre de colonne invalide.")
        total = total * 26 + (ord(ch) - ord('A') + 1)
    return total - 1

def make_unique(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        c = str(c)
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out

def autodetect_name_columns(columns: List[str]) -> Tuple[Optional[str], Optional[str]]:
    lower_map = {c: str(c).strip().lower() for c in columns}
    nom_candidates = [c for c, l in lower_map.items() if any(k in l for k in ["nom", "last name"])]
    prenom_candidates = [c for c, l in lower_map.items() if any(k in l for k in ["prénom", "prenom", "first name"])]
    return (nom_candidates[0] if nom_candidates else None,
            prenom_candidates[0] if prenom_candidates else None)

def autodetect_phone_column(columns: List[str]) -> Optional[str]:
    lower_map = {c: str(c).strip().lower() for c in columns}
    keys = ["téléphone", "telephone", "tel", "phone", "portable", "mobile"]
    for c, l in lower_map.items():
        if any(k in l for k in keys):
            return c
    return None

def autodetect_id_column(columns: List[str]) -> Optional[str]:
    lower_map = {c: str(c).strip().lower() for c in columns}
    for c, l in lower_map.items():
        if re.fullmatch(r".*\bid\b.*", l):
            return c
    return None

def detect_data_start(raw: pd.DataFrame, groupes_col_idx: int, header_row_idx: int) -> int:
    start_probe = header_row_idx + 1
    max_probe = min(len(raw), header_row_idx + 50)
    for r in range(start_probe, max_probe):
        val = raw.iat[r, groupes_col_idx] if groupes_col_idx < raw.shape[1] else None
        if pd.notna(val) and str(val).strip() != "":
            return r
    return start_probe

# --------------------------- Sidebar (commune) ---------------------------
with st.sidebar:
    st.header("⚙️ Import")
    use_sheet = st.text_input("Nom de l'onglet (laisser vide pour auto)", value="")
    col_letter_override = st.text_input("Colonne Groupes (défaut I)", value="I")
    start_row_manual = st.number_input("Forcer ligne de départ (0 = auto)", min_value=0, value=0, step=1)
    show_debug = st.checkbox("Afficher colonnes techniques", value=False)
    st.markdown("---")
    st.header("🧭 Colonnes Nom/Prénom/Téléphone")
    st.caption("Auto-détection, mais tu peux forcer plus bas dans chaque onglet.")
    export_semicolon = st.checkbox("CSV erreurs avec point-virgule (;)", value=True)
    st.caption("Encodage UTF-8-SIG pour Excel FR.")

uploaded = st.file_uploader("Dépose un fichier Excel (.xlsx, .xls)", type=["xlsx", "xls"])
if not uploaded:
    st.info("Charge un fichier pour commencer.")
    st.stop()

xl = pd.ExcelFile(uploaded)
sheet_name = use_sheet if (use_sheet and use_sheet in xl.sheet_names) else xl.sheet_names[0]

try:
    raw = pd.read_excel(uploaded, sheet_name=sheet_name, header=None)
except Exception as e:
    st.error(f"Erreur de lecture: {e}")
    st.stop()

st.write(f"**Onglet lu:** `{sheet_name}`")

# --- I3 / headers / data cut ---
header_row_idx = 2  # I3
try:
    groupes_col_idx = excel_col_to_index(col_letter_override or "I")
except Exception:
    groupes_col_idx = 8  # I
if header_row_idx >= len(raw):
    st.error("La ligne d'en-tête (3) n'existe pas dans ce fichier.")
    st.stop()
if groupes_col_idx >= raw.shape[1]:
    st.error("La colonne Groupes dépasse le nombre de colonnes du fichier.")
    st.stop()

auto_start_row_idx = detect_data_start(raw, groupes_col_idx, header_row_idx)
start_row_idx = int(start_row_manual) - 1 if start_row_manual > 0 else auto_start_row_idx

headers = make_unique(list(raw.iloc[header_row_idx].astype(str)))
data = raw.iloc[start_row_idx:, :].reset_index(drop=True)
if data.shape[1] > len(headers):
    headers = headers + [f"COL_{i}" for i in range(data.shape[1] - len(headers))]
else:
    headers = headers[: data.shape[1]]
data.columns = headers

GROUPES_COL_NAME = "Groupes (détecté I3/auto)"
data[GROUPES_COL_NAME] = raw.iloc[start_row_idx:, groupes_col_idx].reset_index(drop=True)

# Sanity
digits4 = data[GROUPES_COL_NAME].astype(str).str.count(r"\d{4,}").sum()
if digits4 == 0:
    st.warning("⚠️ La colonne **Groupes** semble vide ou mal alignée. "
               "Vérifie l’onglet et/ou force la ligne de départ dans la barre latérale.")
    st.write("Aperçu des 10 premières valeurs de la colonne Groupes :")
    st.write(data[GROUPES_COL_NAME].head(10))

# --------------------------- Onglets ---------------------------
tab_verif, tab_xlsx, tab_pdf = st.tabs(["✅ Vérification", "📄 Listes Excel (1 onglet = 1 classe)", "🖨️ Listes PDF (1 page = 1 classe)"])

# =========================
# Onglet 1 : Vérification
# =========================
with tab_verif:
    st.subheader("Paramètres colonnes (Vérification)")
    nom_guess, prenom_guess = autodetect_name_columns(list(data.columns))
    col1, col2 = st.columns(2)
    with col1:
        nom_col = st.selectbox("Colonne Nom", options=["—"] + list(data.columns),
                               index=(["—"] + list(data.columns)).index(nom_guess) if nom_guess in (["—"] + list(data.columns)) else 0,
                               key="nom_verif")
    with col2:
        prenom_col = st.selectbox("Colonne Prénom", options=["—"] + list(data.columns),
                                  index=(["—"] + list(data.columns)).index(prenom_guess) if prenom_guess in (["—"] + list(data.columns)) else 0,
                                  key="prenom_verif")
    nom_col = None if nom_col == "—" else nom_col
    prenom_col = None if prenom_col == "—" else prenom_col
    if not nom_col or not prenom_col:
        st.warning("⚠️ Choisis/valide les colonnes **Nom** et **Prénom** pour un export d'erreurs correct.")

    # Analyse
    df = data.copy()
    df["Diagnostic"] = df[GROUPES_COL_NAME].apply(analyser_groupes)
    extras = df[GROUPES_COL_NAME].apply(extra_info).apply(pd.Series)
    df = pd.concat([df, extras], axis=1)

    # Répartition
    counts = df["Diagnostic"].value_counts().sort_index()
    total = int(len(df))
    st.markdown(f'<div class="kpi"><b>Total</b><br><span style="font-size:1.4rem">{total}</span></div>', unsafe_allow_html=True)

    st.markdown("#### Répartition par diagnostic")
    rep_df = counts.reset_index()
    rep_df.columns = ["Diagnostic", "Effectif"]
    rep_df.loc[len(rep_df)] = ["Total", total]
    st.dataframe(rep_df, width="stretch")

    # Tableau
    base_cols = [c for c in df.columns if c not in [GROUPES_COL_NAME, "Diagnostic", "FiliereDéduite", "ClasseDéduite", "NumerosTrouvés", "NumerosConnus", "NumerosInconnus"]]
    display_cols = base_cols + [GROUPES_COL_NAME, "Diagnostic"]
    if st.checkbox("Afficher colonnes techniques", value=False, key="tech_verif"):
        display_cols += ["FiliereDéduite", "ClasseDéduite", "NumerosTrouvés", "NumerosConnus", "NumerosInconnus"]
    st.markdown("### Données vérifiées")
    st.dataframe(df[display_cols], width="stretch")

    # Export JSON complet
    records = df.to_dict(orient="records")
    json_bytes = json.dumps(records, ensure_ascii=False, indent=2).encode("utf-8")
    st.download_button("⬇️ Télécharger JSON (complet)", data=json_bytes, file_name="export_verifie.json", mime="application/json", key="json_verif")

    # Export erreurs (Nom, Prénom, Diagnostic) — CSV 3 colonnes
    erreurs = df[df["Diagnostic"] != "OK"].copy()

    def safe_col(s: pd.Series) -> pd.Series:
        return s.astype(str).fillna("").replace({"nan": ""})

    if erreurs.empty:
        st.info("Aucune erreur à exporter 🎉")
    else:
        nom_for_export = nom_col if nom_col in erreurs.columns else None
        prenom_for_export = prenom_col if prenom_col in erreurs.columns else None
        if not nom_for_export:
            erreurs["__NOM__"] = ""
            nom_for_export = "__NOM__"
            st.warning("La colonne Nom sélectionnée n’existe pas — exportera une colonne vide.")
        if not prenom_for_export:
            erreurs["__PRENOM__"] = ""
            prenom_for_export = "__PRENOM__"
            st.warning("La colonne Prénom sélectionnée n’existe pas — exportera une colonne vide.")
        sep = ";" if st.sidebar.checkbox("CSV erreurs avec point-virgule (;)", value=True, key="sep_csv") else ","
        export_df = pd.DataFrame({
            "Nom": safe_col(erreurs[nom_for_export]),
            "Prénom": safe_col(erreurs[prenom_for_export]),
            "Diagnostic": safe_col(erreurs["Diagnostic"]),
        })
        csv_text = export_df.to_csv(index=False, sep=sep)
        csv_bytes = csv_text.encode("utf-8-sig")
        st.download_button("⬇️ Télécharger uniquement les erreurs (CSV) — 3 colonnes", data=csv_bytes,
                           file_name="erreurs_groupes.csv", mime="text/csv", key="csv_erreurs")

# =========================
# Onglet 2 : Excel multi-onglets (1 onglet = 1 classe)
# =========================
with tab_xlsx:
    st.subheader("Paramètres colonnes (Excel)")
    nom_guess2, prenom_guess2 = autodetect_name_columns(list(data.columns))
    tel_guess = autodetect_phone_column(list(data.columns))
    id_guess = autodetect_id_column(list(data.columns))

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        nom_col_x = st.selectbox("Colonne Nom", options=["—"] + list(data.columns),
                                 index=(["—"] + list(data.columns)).index(nom_guess2) if nom_guess2 in (["—"] + list(data.columns)) else 0,
                                 key="nom_xlsx")
    with c2:
        prenom_col_x = st.selectbox("Colonne Prénom", options=["—"] + list(data.columns),
                                    index=(["—"] + list(data.columns)).index(prenom_guess2) if prenom_guess2 in (["—"] + list(data.columns)) else 0,
                                    key="prenom_xlsx")
    with c3:
        tel_col_x = st.selectbox("Colonne Téléphone", options=["—"] + list(data.columns),
                                 index=(["—"] + list(data.columns)).index(tel_guess) if tel_guess in (["—"] + list(data.columns)) else 0,
                                 key="tel_xlsx")
    with c4:
        id_col_x = st.selectbox("Colonne ID (Excel)", options=["—"] + list(data.columns),
                                index=(["—"] + list(data.columns)).index(id_guess) if id_guess in (["—"] + list(data.columns)) else 0,
                                key="id_xlsx")

    nom_col_x = None if nom_col_x == "—" else nom_col_x
    prenom_col_x = None if prenom_col_x == "—" else prenom_col_x
    tel_col_x = None if tel_col_x == "—" else tel_col_x
    id_col_x = None if id_col_x == "—" else id_col_x

    st.markdown("#### Aperçu (10 lignes)")
    st.dataframe(data.head(10), width="stretch")

    # Préparer : classes -> étudiants (ID, Nom, Prénom, Téléphone + Remarque)
    classes_to_students: Dict[int, list] = defaultdict(list)

    def classes_for_row(nums: List[int]) -> Set[int]:
        return {n for n in nums if n in CLASS_NAMES}

    for _, row in data.iterrows():
        nums = parse_numeros(row.get(GROUPES_COL_NAME))
        # EXCLUSION des "numéros exception" dans l'export Excel
        if any(n in EXCEPTION_OK_IF_CLASS_ONLY for n in nums):
            continue
        cls = classes_for_row(nums)
        if not cls:
            continue
        nom_v = "" if not nom_col_x else str(row.get(nom_col_x, "") or "")
        prenom_v = "" if not prenom_col_x else str(row.get(prenom_col_x, "") or "")
        tel_v = "" if not tel_col_x else str(row.get(tel_col_x, "") or "")
        id_v = "" if not id_col_x else str(row.get(id_col_x, "") or "")

        # exclure Salomé Galbois
        if is_salome_galbois(nom_v, prenom_v):
            continue

        for c in cls:
            classes_to_students[c].append((id_v, nom_v, prenom_v, tel_v))

    def sanitize_sheet_name(name: str) -> str:
        # Nettoie pour Excel (<=31 char, pas de : \ / ? * [ ])
        safe = "".join(ch for ch in name if ch not in '[]:*?/\\').strip()
        safe = unicodedata.normalize('NFKD', safe).encode('ascii', 'ignore').decode('ascii')
        return (safe or "Classe")[:31]

    # Choix moteur Excel
    try:
        import xlsxwriter  # noqa: F401
        EXCEL_ENGINE = "xlsxwriter"
    except Exception:
        EXCEL_ENGINE = "openpyxl"

    # Mise en forme selon moteur
    def format_sheet_xlsxwriter(writer, sheet_name, df_len):
        wb = writer.book
        ws = writer.sheets[sheet_name]
        header_fmt = wb.add_format({"bold": True, "bg_color": "#EEEEEE", "border": 1})
        cell_fmt   = wb.add_format({"border": 1})
        widths = [14, 22, 22, 18, 28]  # ID, Nom, Prénom, Téléphone, Remarque
        for col_idx, w in enumerate(widths):
            ws.set_column(col_idx, col_idx, w)
        ws.set_row(0, 18, header_fmt)
        for r in range(1, df_len + 1):
            ws.set_row(r, 16, cell_fmt)

    def idx_to_col(idx: int) -> str:
        s = ""
        idx0 = idx
        while True:
            idx0, r = divmod(idx0, 26)
            s = chr(65 + r) + s
            if idx0 == 0:
                break
            idx0 -= 1
        return s

    def format_sheet_openpyxl(writer, sheet_name, df_len):
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        ws = writer.sheets[sheet_name]
        widths = [14, 22, 22, 18, 28]  # ID, Nom, Prénom, Téléphone, Remarque
        for i, w in enumerate(widths):
            col_letter = idx_to_col(i)
            ws.column_dimensions[col_letter].width = w
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in range(2, df_len + 2):
            for c in range(1, 5 + 1):
                ws.cell(row=r, column=c).border = border

    # Génération Excel
    if st.button("📄 Générer l’Excel (1 onglet = 1 classe)"):
        if not nom_col_x or not prenom_col_x:
            st.error("Sélectionne d'abord **Nom** et **Prénom**.")
        else:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine=EXCEL_ENGINE) as writer:
                if not classes_to_students:
                    df_empty = pd.DataFrame(columns=["ID", "Nom", "Prénom", "Téléphone", "Remarque"])
                    df_empty.to_excel(writer, sheet_name="Aucune classe", index=False)
                    if EXCEL_ENGINE == "xlsxwriter":
                        format_sheet_xlsxwriter(writer, "Aucune classe", 0)
                    else:
                        format_sheet_openpyxl(writer, "Aucune classe", 0)
                else:
                    for ccode in sorted(classes_to_students.keys(), key=lambda c: CLASS_NAMES.get(c, str(c))):
                        label = CLASS_NAMES.get(ccode, f"Classe {ccode}")
                        sheet = sanitize_sheet_name(label)
                        rows_sorted = sorted(classes_to_students[ccode], key=lambda t: ((t[1] or "").lower(), (t[2] or "").lower()))
                        df_sheet = pd.DataFrame(rows_sorted, columns=["ID", "Nom", "Prénom", "Téléphone"])
                        df_sheet["Remarque"] = ""
                        df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                        if EXCEL_ENGINE == "xlsxwriter":
                            format_sheet_xlsxwriter(writer, sheet, len(df_sheet))
                        else:
                            format_sheet_openpyxl(writer, sheet, len(df_sheet))

            buffer.seek(0)
            st.download_button("⬇️ Télécharger l’Excel par classe (.xlsx)", data=buffer,
                               file_name="listes_par_classe.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="xlsx_download")

# =========================
# Onglet 3 : PDF (1 page = 1 classe)
# =========================
with tab_pdf:
    if not REPORTLAB_OK:
        st.error("Le module 'reportlab' n'est pas installé. Ajoute-le à ton environnement pour générer le PDF :\n\npip install reportlab")
    else:
        st.subheader("Paramètres colonnes (PDF)")
        nom_guess3, prenom_guess3 = autodetect_name_columns(list(data.columns))
        tel_guess3 = autodetect_phone_column(list(data.columns))
        id_guess3 = autodetect_id_column(list(data.columns))

        p1, p2, p3, p4 = st.columns(4)
        with p1:
            nom_col_p = st.selectbox("Colonne Nom (PDF)", options=["—"] + list(data.columns),
                                     index=(["—"] + list(data.columns)).index(nom_guess3) if nom_guess3 in (["—"] + list(data.columns)) else 0,
                                     key="nom_pdf")
        with p2:
            prenom_col_p = st.selectbox("Colonne Prénom (PDF)", options=["—"] + list(data.columns),
                                        index=(["—"] + list(data.columns)).index(prenom_guess3) if prenom_guess3 in (["—"] + list(data.columns)) else 0,
                                        key="prenom_pdf")
        with p3:
            tel_col_p = st.selectbox("Colonne Téléphone (PDF)", options=["—"] + list(data.columns),
                                     index=(["—"] + list(data.columns)).index(tel_guess3) if tel_guess3 in (["—"] + list(data.columns)) else 0,
                                     key="tel_pdf")
        with p4:
            id_col_p = st.selectbox("Colonne ID (PDF)", options=["—"] + list(data.columns),
                                    index=(["—"] + list(data.columns)).index(id_guess3) if id_guess3 in (["—"] + list(data.columns)) else 0,
                                    key="id_pdf")

        nom_col_p = None if nom_col_p == "—" else nom_col_p
        prenom_col_p = None if prenom_col_p == "—" else prenom_col_p
        tel_col_p = None if tel_col_p == "—" else tel_col_p
        id_col_p = None if id_col_p == "—" else id_col_p

        # Construire classes->étudiants (mêmes règles d’exclusion)
        classes_to_students_pdf: Dict[int, list] = defaultdict(list)

        def classes_for_row(nums: List[int]) -> Set[int]:
            return {n for n in nums if n in CLASS_NAMES}

        for _, row in data.iterrows():
            nums = parse_numeros(row.get(GROUPES_COL_NAME))
            if any(n in EXCEPTION_OK_IF_CLASS_ONLY for n in nums):
                continue
            cls = classes_for_row(nums)
            if not cls:
                continue
            nom_v = "" if not nom_col_p else str(row.get(nom_col_p, "") or "")
            prenom_v = "" if not prenom_col_p else str(row.get(prenom_col_p, "") or "")
            tel_v = "" if not tel_col_p else str(row.get(tel_col_p, "") or "")
            id_v = "" if not id_col_p else str(row.get(id_col_p, "") or "")

            # exclure Salomé Galbois (PDF aussi)
            if is_salome_galbois(nom_v, prenom_v):
                continue

            for c in cls:
                classes_to_students_pdf[c].append((id_v, nom_v, prenom_v, tel_v))

        st.markdown("#### Aperçu PDF (10 lignes du dataset source)")
        st.dataframe(data.head(10), width="stretch")

        # Génération PDF
        def build_pdf(classes_map: Dict[int, list]) -> bytes:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=A4,
                leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm
            )
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name="ClassTitle", parent=styles["Heading2"], spaceAfter=8, fontSize=14, leading=16))
            styles.add(ParagraphStyle(name="Meta", parent=styles["Normal"], fontSize=8, textColor=colors.grey))

            elements = []
            today = datetime.now().strftime("%d/%m/%Y %H:%M")
            elements.append(Paragraph(f"Généré le {today}", styles["Meta"]))
            elements.append(Spacer(1, 4))

            first = True
            for ccode in sorted(classes_map.keys(), key=lambda c: CLASS_NAMES.get(c, str(c))):
                if not first:
                    elements.append(PageBreak())
                first = False
                label = CLASS_NAMES.get(ccode, f"Classe {ccode}")
                elements.append(Paragraph(label, styles["ClassTitle"]))
                elements.append(Spacer(1, 4))

                # En-têtes PDF : ID, Nom, Prénom, Téléphone, Remarque
                data_tbl = [["ID", "Nom", "Prénom", "Téléphone", "Remarque"]]
                rows_sorted = sorted(classes_map[ccode], key=lambda t: ((t[1] or "").lower(), (t[2] or "").lower()))
                for id_v, nom_v, prenom_v, tel_v in rows_sorted:
                    data_tbl.append([id_v, nom_v, prenom_v, tel_v, ""])

                col_widths = [18*mm, 45*mm, 45*mm, 30*mm, 42*mm]

                tbl = Table(data_tbl, colWidths=col_widths, hAlign="LEFT")
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),

                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),

                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]))
                elements.append(tbl)
                elements.append(Spacer(1, 6))
                elements.append(Paragraph("<i>Chaque classe commence sur une nouvelle page.</i>", styles["Meta"]))

            doc.build(elements)
            buffer.seek(0)
            return buffer.getvalue()

        # Bouton PDF
        if st.button("🖨️ Générer le PDF (1 page = 1 classe)"):
            if not nom_col_p or not prenom_col_p:
                st.error("Sélectionne d'abord **Nom** et **Prénom**.")
            else:
                pdf_bytes = build_pdf(classes_to_students_pdf if classes_to_students_pdf else {})
                st.download_button(
                    "⬇️ Télécharger le PDF par classe",
                    data=pdf_bytes,
                    file_name="listes_par_classe.pdf",
                    mime="application/pdf",
                    key="pdf_download"
                )
